

# Project Scrape From Google
# Scrape From Google numbers and names

## To Do List:
#

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import datetime
import pandas as pd
import pandasgui as pg
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def make_driver():   # פונקציית הגדרת הדרייבר והוספת אפשרויות לדפדפן
    PATH = "C:\Program Files (x86)\chromedriver.exe"  # מיקום הפאטצ'
    service = Service(PATH)  # הפעלת הפאטצ'
    options = webdriver.ChromeOptions()  # כדי שהחלון של הדפדפן לא ייסגר ישר
    options.add_experimental_option("detach", True)  # כדי שהחלון של הדפדפן לא ייסגר ישר
    return webdriver.Chrome(service=service, options=options)  # הגדרת הדרייבר שאיתו נשתמש

def save_and_font(path, file_name):
    david_font = Font(size=11, name='David')  # Apply the new font style to all cells in the sheet
    alignment = Alignment(vertical='center', horizontal='center')  # הגדרות לתא שהכתב יהיה במרכז התא

    for row in ws.rows:   # מעבר על כל שורה ואז על כל תא בכל שורה ועדכון הפונט וההגדרות של הכתב
        for cell in row:
            cell.font = david_font
            cell.alignment = alignment

    for i in range(1, 5):   # הדגשת 4 השורות הראשונות
        for cell in ws[i]:
            cell.font = Font(bold=True, size=11, name='David')

    wb.save(path + f'{file_name}.xlsx')   # שמירת קובץ האקסל עם שם ומיקום לפי הקלט
    return

def sheet_structure(ws):  # יצירת מבנה ושלד המסמך
    ws.sheet_view.rightToLeft = True   # הפיכת המסמך לכתב מימין לשמאל

    ws['d2'] = 'הנושא - פרטים '
    ws.merge_cells('d2:f2')
    ws['a4'] = 'שם: '
    ws['b4'] = 'מספר טלפון: '
    ws['c4'] = 'כתובת: '
    ws['d4'] = 'אתר אינטרנט: '
    ws['g4'] = ' עוד: '

    ws.column_dimensions["a"].width = len(ws.cell(4, 1).value)+23  # עדכון רוחב העמודות לפי הכתוב בכותרת ועוד קצת לכל עמודה
    ws.column_dimensions["b"].width = len(ws.cell(4, 2).value)+2
    ws.column_dimensions["c"].width = len(ws.cell(4, 3).value)+14
    ws.column_dimensions["d"].width = len(ws.cell(4, 4).value)+12
    ws.column_dimensions["g"].width = len(ws.cell(4, 7).value)+3
    ws.freeze_panes = 'A5'                                        # הקפאת השורות 1-4
    return

def write_in_the_exel(new_df, new_row):   # מוסיף לאקסל את הרשימה עם הפרטים של כל מודעה
    ws.cell(len(new_df)+4, 1, new_row[0])  # כותב באקסל בשורה של האינדיקטור ורץ על העמודות להזין את כל הנתונים
    ws.cell(len(new_df)+4, 2, new_row[1])
    ws.cell(len(new_df)+4, 3, new_row[2])
    ws.cell(len(new_df)+4, 4, new_row[3])
    return


#  The Main Program:

wb = Workbook()   # יצירת מסמך חדש
ws = wb.active    # הגדרת הטאב שעליו נעבוד
sheet_structure(ws)

new_df = pd.DataFrame({"שם ": [], 'מספר טלפון': [], "כתובת": [], "אתר אינטרנט": []})  # הגדרת התגיות לDF בפנדה
driver = make_driver()
driver.get("https://www.google.com/search?q=..................")  # כתובת האתר שאני רוצה לגשת אליו ולחפש בו
time.sleep(2)

while True:        # כביכול לולאה אינסופית (עד לברייק כאשר יש שגיאה בטעינה העמוד הבא)
    element = driver.find_elements(By.CSS_SELECTOR, '[jscontroller="AtSb"]')   # מוציא את כל האלמנטים/ השורות של כל המודעות
    for elm in element:   # הלולאה רצה מודעה מסויימת כל פעם
        new_row = []   # רשימה שתתאפס כל מודעה, ועליה אשמור את הפרטים שארצה לשמור
        row = elm.find_element(By.CLASS_NAME, "cXedhc")  # מאתר חלק במודעה (שאפשר ללחוץ עליו)
        row.click()    # לוחץ כדי לפתוח את פרטי
        time.sleep(2)
        name = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.Ftghae')))  # מאתר את אלמנט השם (חייבים להשתמש בWAIT כדי שהשם ייטען)
        new_row.append(name.text)    # הוספת הטקסט של השם לרשימה
        print(name.text)
        try:
            phone_number = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "LrzXr.zdqRlf.kno-fv")))  # מאתר את האלמנט של מספר הטלפון
            new_row.append(phone_number.text) # הוספת טקסט מספר הטלפון לרשימה
            print(phone_number.text)
        except:
            phone_number = "אין מספר טלפון זמין "  # מכיוון ולא לכל אחד יש מספר טלפון זמין בגוגל, אז משתמש בTRY-EXCEPT
            new_row.append(phone_number)   # ומוסיף ערך לרשימה כדי שלא יבלגן את שאר הסדר (כי אני משתמש באפפנד ולא לפי אינדקס)
            print(phone_number)
        try:
            address = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "LrzXr")))  # מאתר את האלמנט של הכתובת
            new_row.append(address.text)   # הוספת הטקסט של כתובת המשרד
            print(address.text)
        except:
            address = "אין כתובת זמינה "  # אותו כנ"ל כמו מעל
            new_row.append(address)
            print(address)
        try:
            a_tag = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "dHS6jb")))   # מאתר את כתובת האינטרנט
            link = a_tag.get_attribute("href")        # מוציא מהאלמנט את הקישור המדוייק
            new_row.append(link)                      # מוסיף את הקישור לרשימה
            print(link)
        except:
            link = "אין אתר אינטרנט זמין "
            new_row.append(link)
            print(link)

        new_df.loc[len(new_df)] = new_row  # הוספת השורה לטבלה
        write_in_the_exel(new_df, new_row)    # כתיבת שורה באקסל על ידי הרשימה עם הפרטים המלאים (לפני שהרשימה מתאפסת)
    pg.show(new_df)       # הצגה חיצונית של הDF
    print(new_df)
    try:
        next_page_tag = driver.find_element(By.ID, "pnnext")   # מוצא את האלמנט של העמוד הבא
        next_page_link = next_page_tag.get_attribute("href")  # מוציא את הלינק לעמוד הבא
        print(next_page_link)
    except Exception as e:
        break                     # עצירת הלולאה - ברגע שלא מוצאים עוד אלמנט של עמוד הבא אז עולה שגיאה, ויוצאים מהלולאה הראשית
    # Store the current window handle
    main_window = driver.current_window_handle  # שומר את העמוד הנוכחי
    driver.execute_script("window.open('');")   # פותח טאב חדש ריק
    driver.switch_to.window(driver.window_handles[-1]) # מעביר את הדריבר לחלון החדש
    driver.get(next_page_link)   # ומכניס את הלינק של העמוד הבא שהוצאתי, ומריץ את הדרייבר
    time.sleep(2)

print(new_df[new_df['A'].duplicated()])   # לוקח רק את הטבלה שבה יש כפילויות בשם
print(new_df[new_df['B'].duplicated()])   # לוקח רק את הטבלה שבה יש כפילויות במספר הטלפון

file_name = 'your name for the file'
path = 'C:/Users/........../'
save_and_font(path, file_name)




